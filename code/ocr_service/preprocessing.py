"""
Document Preprocessing Module

This module handles all document preprocessing operations including:
- Deskewing
- Noise reduction
- Contrast enhancement
- Binarization
- Image sharpening
- Morphological operations
"""

import cv2
import numpy as np
import logging
from typing import Tuple, Optional, Dict, Any
from PIL import Image, ImageEnhance, ImageFilter
import math


logger = logging.getLogger(__name__)


class DocumentPreprocessor:
    """Handles document preprocessing for OCR optimization."""
    
    def __init__(self, config: Dict[str, Any]):
        """
        Initialize preprocessor with configuration.
        
        Args:
            config: Preprocessing configuration dictionary
        """
        self.config = config
        self.logger = logging.getLogger(__name__)
    
    def preprocess_image(self, image: np.ndarray) -> np.ndarray:
        """
        Apply complete preprocessing pipeline to an image.
        
        Args:
            image: Input image as numpy array
            
        Returns:
            Preprocessed image
        """
        try:
            # Ensure image is in grayscale for better OCR
            if len(image.shape) == 3:
                image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            # Step 1: Noise reduction
            image = self.remove_noise(image)
            
            # Step 2: Deskewing
            image = self.deskew_image(image)
            
            # Step 3: Contrast enhancement
            image = self.enhance_contrast(image)
            
            # Step 4: Sharpening
            image = self.sharpen_image(image)
            
            # Step 5: Binarization
            image = self.binarize_image(image)
            
            return image
            
        except Exception as e:
            self.logger.error(f"Error in preprocessing pipeline: {str(e)}")
            raise
    
    def remove_noise(self, image: np.ndarray) -> np.ndarray:
        """
        Remove noise from the image.
        
        Args:
            image: Input grayscale image
            
        Returns:
            Denoised image
        """
        try:
            # Apply Gaussian blur for noise reduction
            ksize = self.config.get('gaussian_blur_ksize', 3)
            if ksize % 2 == 0:
                ksize += 1  # Ensure odd kernel size
            
            denoised = cv2.GaussianBlur(image, (ksize, ksize), 0)
            
            # Apply median filter for salt-and-pepper noise
            denoised = cv2.medianBlur(denoised, 3)
            
            self.logger.debug("Noise reduction applied successfully")
            return denoised
            
        except Exception as e:
            self.logger.error(f"Error in noise removal: {str(e)}")
            return image
    
    def deskew_image(self, image: np.ndarray) -> np.ndarray:
        """
        Correct skew in the image.
        
        Args:
            image: Input grayscale image
            
        Returns:
            Deskewed image
        """
        try:
            # Detect skew angle using Hough transform
            edges = cv2.Canny(image, 50, 150, apertureSize=3)
            lines = cv2.HoughLines(edges, 1, np.pi / 180, threshold=100)
            
            if lines is not None:
                angles = []
                for rho, theta in lines[0]:
                    angle = theta * 180 / np.pi
                    if 0 < angle < 180:
                        angles.append(angle - 90)
                
                if angles:
                    median_angle = np.median(angles)
                    angle_threshold = self.config.get('deskew_angle_threshold', 0.5)
                    max_angle = self.config.get('deskew_max_angle', 5.0)
                    
                    if abs(median_angle) > angle_threshold and abs(median_angle) < max_angle:
                        # Rotate image to correct skew
                        rows, cols = image.shape
                        rotation_matrix = cv2.getRotationMatrix2D(
                            (cols / 2, rows / 2), median_angle, 1.0
                        )
                        deskewed = cv2.warpAffine(
                            image, rotation_matrix, (cols, rows),
                            flags=cv2.INTER_LINEAR, borderMode=cv2.BORDER_REPLICATE
                        )
                        
                        self.logger.debug(f"Image deskewed by {median_angle:.2f} degrees")
                        return deskewed
            
            return image
            
        except Exception as e:
            self.logger.error(f"Error in deskewing: {str(e)}")
            return image
    
    def enhance_contrast(self, image: np.ndarray) -> np.ndarray:
        """
        Enhance image contrast.
        
        Args:
            image: Input grayscale image
            
        Returns:
            Contrast-enhanced image
        """
        try:
            # Apply histogram equalization
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
            enhanced = clahe.apply(image)
            
            # Apply gamma correction if configured
            gamma = self.config.get('gamma_correction', 1.0)
            if gamma != 1.0:
                inv_gamma = 1.0 / gamma
                table = np.array([(i / 255.0) ** inv_gamma * 255 for i in range(256)]).astype('uint8')
                enhanced = cv2.LUT(enhanced, table)
            
            # Apply brightness and contrast adjustments
            contrast_factor = self.config.get('contrast_factor', 1.5)
            brightness_delta = self.config.get('brightness_delta', 10)
            
            enhanced = cv2.convertScaleAbs(
                enhanced, 
                alpha=contrast_factor, 
                beta=brightness_delta
            )
            
            self.logger.debug("Contrast enhancement applied successfully")
            return enhanced
            
        except Exception as e:
            self.logger.error(f"Error in contrast enhancement: {str(e)}")
            return image
    
    def sharpen_image(self, image: np.ndarray) -> np.ndarray:
        """
        Sharpen the image.
        
        Args:
            image: Input grayscale image
            
        Returns:
            Sharpened image
        """
        try:
            kernel = self.config.get('sharpen_kernel')
            if kernel:
                sharpened = cv2.filter2D(image, -1, np.array(kernel))
                self.logger.debug("Image sharpening applied successfully")
                return sharpened
            
            return image
            
        except Exception as e:
            self.logger.error(f"Error in image sharpening: {str(e)}")
            return image
    
    def binarize_image(self, image: np.ndarray) -> np.ndarray:
        """
        Apply binarization to the image.
        
        Args:
            image: Input grayscale image
            
        Returns:
            Binarized image
        """
        try:
            method = self.config.get('binarization_method', 'adaptive')
            
            if method == 'simple':
                threshold = self.config.get('binarization_threshold', 127)
                _, binary = cv2.threshold(image, threshold, 255, cv2.THRESH_BINARY)
                
            elif method == 'adaptive':
                binary = cv2.adaptiveThreshold(
                    image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                    cv2.THRESH_BINARY, 11, 2
                )
                
            elif method == 'otsu':
                _, binary = cv2.threshold(image, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                
            else:
                binary = image
            
            # Apply morphological operations
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
            binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
            
            self.logger.debug(f"Binarization method '{method}' applied successfully")
            return binary
            
        except Exception as e:
            self.logger.error(f"Error in binarization: {str(e)}")
            return image
    
    def normalize_image_size(self, image: np.ndarray, max_size: Tuple[int, int] = (2000, 2000)) -> np.ndarray:
        """
        Normalize image size while maintaining aspect ratio.
        
        Args:
            image: Input image
            max_size: Maximum (width, height)
            
        Returns:
            Resized image
        """
        try:
            height, width = image.shape[:2]
            max_width, max_height = max_size
            
            if width > max_width or height > max_height:
                scale = min(max_width / width, max_height / height)
                new_width = int(width * scale)
                new_height = int(height * scale)
                
                resized = cv2.resize(image, (new_width, new_height), interpolation=cv2.INTER_AREA)
                self.logger.debug(f"Image resized from {width}x{height} to {new_width}x{new_height}")
                return resized
            
            return image
            
        except Exception as e:
            self.logger.error(f"Error in size normalization: {str(e)}")
            return image
    
    def enhance_resolution(self, image: np.ndarray, scale_factor: int = 2) -> np.ndarray:
        """
        Enhance image resolution using interpolation.
        
        Args:
            image: Input image
            scale_factor: Factor by which to increase resolution
            
        Returns:
            Enhanced resolution image
        """
        try:
            height, width = image.shape[:2]
            new_width = width * scale_factor
            new_height = height * scale_factor
            
            enhanced = cv2.resize(image, (new_width, new_height), interpolation=cv2.INTER_CUBIC)
            self.logger.debug(f"Resolution enhanced by factor {scale_factor}")
            return enhanced
            
        except Exception as e:
            self.logger.error(f"Error in resolution enhancement: {str(e)}")
            return image
    
    def remove_shadows(self, image: np.ndarray) -> np.ndarray:
        """
        Remove shadows from the image.
        
        Args:
            image: Input grayscale image
            
        Returns:
            Image with shadows removed
        """
        try:
            # Apply morphological operations to remove shadows
            kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
            opened = cv2.morphologyEx(image, cv2.MORPH_OPEN, kernel)
            
            # Subtract opened image from original to get shadow regions
            shadow_mask = cv2.subtract(image, opened)
            
            # Dilate shadow mask to include shadow areas
            shadow_mask = cv2.dilate(shadow_mask, kernel, iterations=2)
            
            # Inpaint shadow areas
            inpainted = cv2.inpaint(image, shadow_mask, 3, cv2.INPAINT_TELEA)
            
            self.logger.debug("Shadow removal applied successfully")
            return inpainted
            
        except Exception as e:
            self.logger.error(f"Error in shadow removal: {str(e)}")
            return image
    
    def detect_edges(self, image: np.ndarray) -> np.ndarray:
        """
        Detect edges in the image.
        
        Args:
            image: Input grayscale image
            
        Returns:
            Edge-detected image
        """
        try:
            # Apply Canny edge detection
            edges = cv2.Canny(image, 50, 150)
            
            # Dilate edges to make them more prominent
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
            edges = cv2.dilate(edges, kernel)
            
            self.logger.debug("Edge detection applied successfully")
            return edges
            
        except Exception as e:
            self.logger.error(f"Error in edge detection: {str(e)}")
            return image
    
    def extract_text_regions(self, image: np.ndarray) -> list:
        """
        Extract potential text regions from the image.
        
        Args:
            image: Input image
            
        Returns:
            List of text region bounding boxes
        """
        try:
            # Apply threshold to get binary image
            if len(image.shape) == 3:
                gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            else:
                gray = image.copy()
            
            _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # Find contours
            contours, _ = cv2.findContours(
                binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
            )
            
            # Filter contours by size and aspect ratio
            text_regions = []
            for contour in contours:
                x, y, w, h = cv2.boundingRect(contour)
                
                # Filter based on size
                if w > 30 and h > 30 and w < image.shape[1] * 0.9:
                    aspect_ratio = w / float(h)
                    
                    # Filter based on aspect ratio (typical text regions)
                    if 0.5 < aspect_ratio < 5.0:
                        text_regions.append((x, y, w, h))
            
            self.logger.debug(f"Extracted {len(text_regions)} potential text regions")
            return text_regions
            
        except Exception as e:
            self.logger.error(f"Error in text region extraction: {str(e)}")
            return []
    
    def crop_text_regions(self, image: np.ndarray, regions: list) -> list:
        """
        Crop text regions from the image.
        
        Args:
            image: Input image
            regions: List of (x, y, w, h) tuples
            
        Returns:
            List of cropped text region images
        """
        cropped_images = []
        for x, y, w, h in regions:
            try:
                cropped = image[y:y+h, x:x+w]
                cropped_images.append(cropped)
            except Exception as e:
                self.logger.error(f"Error cropping region ({x}, {y}, {w}, {h}): {str(e)}")
                continue
        
        self.logger.debug(f"Cropped {len(cropped_images)} text regions")
        return cropped_images


class PDFProcessor:
    """Handles PDF document preprocessing."""
    
    def __init__(self, config: Dict[str, Any]):
        """
        Initialize PDF processor.
        
        Args:
            config: PDF processing configuration
        """
        self.config = config
        self.logger = logging.getLogger(__name__)
    
    def extract_images_from_pdf(self, pdf_path: str) -> list:
        """
        Extract images from PDF pages.
        
        Args:
            pdf_path: Path to PDF file
            
        Returns:
            List of extracted images
        """
        try:
            import fitz  # PyMuPDF
            
            doc = fitz.open(pdf_path)
            images = []
            
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                
                # Get page as image
                mat = self.config.get('dpi', 300) / 72  # Convert DPI to matrix
                pix = page.get_pixmap(matrix=fitz.Matrix(mat, mat))
                
                # Convert to numpy array
                img_data = pix.tobytes("png")
                img_array = np.frombuffer(img_data, np.uint8)
                img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
                
                if img is not None:
                    images.append(img)
                    self.logger.debug(f"Extracted image from PDF page {page_num + 1}")
            
            doc.close()
            self.logger.info(f"Extracted {len(images)} images from PDF")
            return images
            
        except ImportError:
            raise ImportError("PyMuPDF (fitz) is required for PDF processing")
        except Exception as e:
            self.logger.error(f"Error extracting images from PDF: {str(e)}")
            raise


class SpreadsheetProcessor:
    """Handles spreadsheet document preprocessing."""
    
    def __init__(self, config: Dict[str, Any]):
        """
        Initialize spreadsheet processor.
        
        Args:
            config: Spreadsheet processing configuration
        """
        self.config = config
        self.logger = logging.getLogger(__name__)
    
    def extract_text_from_spreadsheet(self, file_path: str) -> str:
        """
        Extract text from spreadsheet files.
        
        Args:
            file_path: Path to spreadsheet file
            
        Returns:
            Extracted text content
        """
        try:
            ext = file_path.split('.')[-1].lower()
            
            if ext in ['xlsx', 'xls']:
                return self._extract_from_excel(file_path)
            elif ext == 'csv':
                return self._extract_from_csv(file_path)
            elif ext == 'ods':
                return self._extract_from_ods(file_path)
            else:
                raise ValueError(f"Unsupported spreadsheet format: {ext}")
                
        except Exception as e:
            self.logger.error(f"Error extracting text from spreadsheet: {str(e)}")
            raise
    
    def _extract_from_excel(self, file_path: str) -> str:
        """Extract text from Excel files."""
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text_content = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content.append(f"\n=== Sheet: {sheet_name} ===\n")
                
                for row in sheet.iter_rows():
                    row_text = []
                    for cell in row:
                        if cell.value:
                            row_text.append(str(cell.value))
                    if row_text:
                        text_content.append("\t".join(row_text))
            
            return "\n".join(text_content)
            
        except ImportError:
            raise ImportError("openpyxl is required for Excel processing")
    
    def _extract_from_csv(self, file_path: str) -> str:
        """Extract text from CSV files."""
        try:
            import pandas as pd
            
            df = pd.read_csv(file_path)
            return df.to_string(index=False)
            
        except ImportError:
            raise ImportError("pandas is required for CSV processing")
    
    def _extract_from_ods(self, file_path: str) -> str:
        """Extract text from ODS files."""
        try:
            import pandas as pd
            
            df = pd.read_excel(file_path, engine='odf')
            return df.to_string(index=False)
            
        except ImportError:
            raise ImportError("pandas and odfpy are required for ODS processing")
